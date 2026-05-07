"""
sgi_interface.py
---------------
Graphical interface for the SGI data extractor.
Run from the same folder as extract_sgi.py.

Requires:
    pip install python-docx openpyxl pillow
"""

import os
import sys
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path

# ── Try to load Pillow to display the PNG logo ──────────────────────────────
try:
    from PIL import Image, ImageTk
    PILLOW_OK = True
except ImportError:
    PILLOW_OK = False


# ────────────────────────────────────────────────────────────────────────────
# Color Palette
# ────────────────────────────────────────────────────────────────────────────
HEADER_BLUE   = "#0D3B6E"   # Institutional dark blue
BUTTON_BLUE   = "#1565C0"   # Active button blue
HOVER_BLUE    = "#1976D2"   # Button hover blue
DISABLED_GRAY = "#90A4AE"   # Disabled button gray
OK_GREEN      = "#2E7D32"   # Completion bar green
RUNNING_GRAY  = "#546E7A"   # Execution bar gray
WHITE         = "#FFFFFF"
BODY_BG       = "#FFFFFF"
FIELD_BG      = "#F5F7FA"
FIELD_BORDER  = "#CFD8DC"
DARK_TEXT     = "#1A237E"
LABEL_TEXT    = "#37474F"
SUBHEADER_BG  = "#E3EBF6"


# ────────────────────────────────────────────────────────────────────────────
# Main Class
# ────────────────────────────────────────────────────────────────────────────
class SGIApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("SGC Data Extractor")
        self.resizable(False, False)
        self.configure(bg=BODY_BG)

        # Try to load .ico icon
        ico_path = Path("../img/tecnm.ico")
        if ico_path.exists():
            try:
                self.iconbitmap(str(ico_path))
            except Exception:
                pass

        # Variables
        self.source_folder_var = tk.StringVar()
        self.dest_file_var = tk.StringVar()
        self.is_running = False

        self._build_ui()
        self._center_window()

    # ── Center window on screen ─────────────────────────────────────────────
    def _center_window(self):
        self.update_idletasks()
        w = self.winfo_width()
        h = self.winfo_height()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.geometry(f"+{x}+{y}")

    # ── UI Construction ─────────────────────────────────────────────────────
    def _build_ui(self):
        self._build_header()
        self._build_subheader()
        self._build_body()

    # ── Blue Header with Logo ───────────────────────────────────────────────
    def _build_header(self):
        header = tk.Frame(self, bg=HEADER_BLUE, height=80)
        header.pack(fill="x")
        header.pack_propagate(False)

        # PNG Logo on the left
        logo_path = Path("../img/tecnm.png")
        if PILLOW_OK and logo_path.exists():
            try:
                img = Image.open(logo_path).convert("RGBA")
                img.thumbnail((60, 60), Image.LANCZOS)
                self._logo_img = ImageTk.PhotoImage(img)
                tk.Label(
                    header, image=self._logo_img,
                    bg=HEADER_BLUE, bd=0
                ).pack(side="left", padx=(18, 10), pady=10)
            except Exception:
                self._logo_img = None
        else:
            # Placeholder square if no image
            placeholder = tk.Label(
                header, text="TecNM", bg=WHITE, fg=HEADER_BLUE,
                font=("Georgia", 9, "bold"), width=7, height=3, relief="flat"
            )
            placeholder.pack(side="left", padx=(18, 10), pady=10)

        # Header Text
        text_frame = tk.Frame(header, bg=HEADER_BLUE)
        text_frame.pack(side="left", fill="both", expand=True, pady=10)

        tk.Label(
            text_frame,
            text="Tecnológico Nacional de México",
            font=("Georgia", 15, "bold"),
            bg=HEADER_BLUE, fg=WHITE, anchor="w"
        ).pack(anchor="w")

        tk.Label(
            text_frame,
            text="Campus San Juan del Río",
            font=("Georgia", 11),
            bg=HEADER_BLUE, fg="#BBDEFB", anchor="w"
        ).pack(anchor="w")

    # ── Institutional Subheader ─────────────────────────────────────────────
    def _build_subheader(self):
        sub = tk.Frame(self, bg=SUBHEADER_BG, pady=8)
        sub.pack(fill="x")

        tk.Label(
            sub,
            text="Industrial Engineering Career Department",
            font=("Segoe UI", 10, "bold"),
            bg=SUBHEADER_BG, fg=DARK_TEXT
        ).pack()

        tk.Label(
            sub,
            text="Supervisor:  María Patricia Uribe Rodríguez",
            font=("Segoe UI", 9),
            bg=SUBHEADER_BG, fg=LABEL_TEXT
        ).pack()

    # ── Main Body ───────────────────────────────────────────────────────────
    def _build_body(self):
        body = tk.Frame(self, bg=BODY_BG, padx=32, pady=24)
        body.pack(fill="both", expand=True)

        # Program Title
        tk.Label(
            body,
            text="SGC  —  Data Extractor",
            font=("Georgia", 14, "bold"),
            bg=BODY_BG, fg=DARK_TEXT
        ).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 20))

        # ── Source Folder Selector ─────────────────────────────────────────
        self._build_field_row(
            body, row=1,
            label="ZIP / DOCX Files Folder:",
            variable=self.source_folder_var,
            btn_text="📂  Select Folder",
            cmd=self._select_folder,
            placeholder="Select the folder containing the files..."
        )

        # ── Destination Excel File Selector ────────────────────────────────
        self._build_field_row(
            body, row=3,
            label="Save Excel to:",
            variable=self.dest_file_var,
            btn_text="💾  Save As...",
            cmd=self._select_destination,
            placeholder="Choose where to save the data.xlsx file..."
        )

        # Separator
        sep = tk.Frame(body, bg=FIELD_BORDER, height=1)
        sep.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(20, 0))

        # ── Start Button ────────────────────────────────────────────────────
        self.start_btn = tk.Button(
            body,
            text="▶   Start Extraction",
            font=("Segoe UI", 11, "bold"),
            bg=DISABLED_GRAY, fg=WHITE,
            activebackground=HOVER_BLUE, activeforeground=WHITE,
            relief="flat", cursor="arrow",
            padx=28, pady=10,
            state="disabled",
            command=self._start_extraction
        )
        self.start_btn.grid(row=6, column=0, columnspan=3, pady=(18, 0))

        # ── Status Bar ──────────────────────────────────────────────────────
        status_frame = tk.Frame(body, bg=BODY_BG)
        status_frame.grid(row=7, column=0, columnspan=3, sticky="ew", pady=(20, 0))
        status_frame.columnconfigure(0, weight=1)

        self.status_lbl = tk.Label(
            status_frame,
            text="",
            font=("Segoe UI", 9),
            bg=BODY_BG, fg=LABEL_TEXT, anchor="w"
        )
        self.status_lbl.grid(row=0, column=0, sticky="w", pady=(0, 4))

        # Progress Bar (Custom Style)
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure(
            "idle.Horizontal.TProgressbar",
            troughcolor=FIELD_BORDER, background=FIELD_BORDER,
            thickness=14, borderwidth=0
        )
        style.configure(
            "running.Horizontal.TProgressbar",
            troughcolor=FIELD_BORDER, background=RUNNING_GRAY,
            thickness=14, borderwidth=0
        )
        style.configure(
            "done.Horizontal.TProgressbar",
            troughcolor=FIELD_BORDER, background=OK_GREEN,
            thickness=14, borderwidth=0
        )

        self.progressbar = ttk.Progressbar(
            status_frame,
            style="idle.Horizontal.TProgressbar",
            mode="indeterminate", length=460
        )
        self.progressbar.grid(row=1, column=0, sticky="ew")

        body.columnconfigure(1, weight=1)
        body.config(width=560)

    # ── Helper: Label + Entry + Button Row ─────────────────────────────────
    def _build_field_row(self, parent, row, label, variable, btn_text, cmd, placeholder):
        tk.Label(
            parent, text=label,
            font=("Segoe UI", 9, "bold"),
            bg=BODY_BG, fg=LABEL_TEXT, anchor="w"
        ).grid(row=row, column=0, columnspan=3, sticky="w", pady=(0, 4))

        entry_frame = tk.Frame(parent, bg=FIELD_BORDER, bd=0)
        entry_frame.grid(row=row+1, column=0, columnspan=2, sticky="ew", ipady=1, ipadx=1)
        entry_frame.columnconfigure(0, weight=1)

        entry = tk.Entry(
            entry_frame,
            textvariable=variable,
            font=("Segoe UI", 9),
            bg=FIELD_BG, fg=LABEL_TEXT,
            relief="flat", state="readonly",
            readonlybackground=FIELD_BG,
            bd=6, width=48
        )
        entry.grid(row=0, column=0, sticky="ew")

        # Visual Placeholder
        if not variable.get():
            entry.config(state="normal")
            entry.insert(0, placeholder)
            entry.config(fg="#B0BEC5", state="readonly")

        def _on_var_change(*_):
            val = variable.get()
            entry.config(state="normal")
            entry.delete(0, "end")
            if val:
                entry.insert(0, val)
                entry.config(fg=LABEL_TEXT)
            else:
                entry.insert(0, placeholder)
                entry.config(fg="#B0BEC5")
            entry.config(state="readonly")
            self._update_start_button()

        variable.trace_add("write", _on_var_change)

        # Store entry reference
        attr = "folder_entry" if "folder" in label.lower() else "dest_entry"
        setattr(self, attr, entry)

        btn = tk.Button(
            parent, text=btn_text,
            font=("Segoe UI", 9),
            bg=BUTTON_BLUE, fg=WHITE,
            activebackground=HOVER_BLUE, activeforeground=WHITE,
            relief="flat", cursor="hand2",
            padx=12, pady=6,
            command=cmd
        )
        btn.grid(row=row+1, column=2, padx=(10, 0), sticky="w")
        self._add_hover(btn, BUTTON_BLUE, HOVER_BLUE)

        # Store button reference
        attr_btn = "folder_btn" if "folder" in label.lower() else "dest_btn"
        setattr(self, attr_btn, btn)

        parent.rowconfigure(row+2, minsize=12)

    # ── Hover Helper ────────────────────────────────────────────────────────
    @staticmethod
    def _add_hover(widget, normal_color, hover_color):
        widget.bind("<Enter>", lambda e: widget.config(bg=hover_color))
        widget.bind("<Leave>", lambda e: widget.config(bg=normal_color))

    # ── Source Folder Selection ────────────────────────────────────────────
    def _select_folder(self):
        folder = filedialog.askdirectory(
            title="Select Folder with ZIP / DOCX Files",
            mustexist=True
        )
        if folder:
            self.source_folder_var.set(folder)

    # ── Destination File Selection ─────────────────────────────────────────
    def _select_destination(self):
        file = filedialog.asksaveasfilename(
            title="Save Excel As...",
            defaultextension=".xlsx",
            filetypes=[("Excel File", "*.xlsx"), ("All Files", "*.*")],
            initialfile="data.xlsx"
        )
        if file:
            self.dest_file_var.set(file)

    # ── Enable / Disable Start Button ──────────────────────────────────────
    def _update_start_button(self):
        if self.is_running:
            return
        if self.source_folder_var.get() and self.dest_file_var.get():
            self.start_btn.config(
                state="normal", bg=BUTTON_BLUE, cursor="hand2"
            )
            self._add_hover(self.start_btn, BUTTON_BLUE, HOVER_BLUE)
        else:
            self.start_btn.config(
                state="disabled", bg=DISABLED_GRAY, cursor="arrow"
            )

    # ── Lock / Unlock Controls ─────────────────────────────────────────────
    def _set_controls(self, locked: bool):
        state = "disabled" if locked else "normal"
        self.folder_btn.config(state=state)
        self.dest_btn.config(state=state)
        if not locked:
            self._update_start_button()
        else:
            self.start_btn.config(state="disabled", bg=DISABLED_GRAY, cursor="arrow")

    # ── Start Extraction in Separate Thread ────────────────────────────────
    def _start_extraction(self):
        if self.is_running:
            return
        self.is_running = True
        self._set_controls(locked=True)

        # Status: Running
        self.progressbar.config(style="running.Horizontal.TProgressbar", mode="indeterminate")
        self.progressbar.start(12)
        self.status_lbl.config(text="⏳  Extracting data, please wait...", fg=RUNNING_GRAY)

        thread = threading.Thread(target=self._run_extraction, daemon=True)
        thread.start()

    def _run_extraction(self):
        """Runs extract_sgi in a thread to avoid freezing the UI."""
        try:
            # Import the extraction module (must be in the same folder)
            import importlib.util

            spec = importlib.util.spec_from_file_location(
                "sgi_extract",
                Path(__file__).parent / "sgi_extract.py"
            )
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)

            folder = self.source_folder_var.get()
            destination = self.dest_file_var.get()

            # ── Capture stdout to avoid breaking UI ─────────────────────
            import io, contextlib
            buffer = io.StringIO()
            data_list = [mod.EXCEL_HEADER]

            with contextlib.redirect_stdout(buffer):
                mod.process_path(folder, data_list)

            log_output = buffer.getvalue()

            if len(data_list) > 1:
                # Save to the user-selected path
                os.chdir(str(Path(destination).parent))

                # Rewrite Excel with correct styling
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "SGI Data"

                header_fill = PatternFill("solid", fgColor="1F4E79")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

                for col_idx, field in enumerate(data_list[0], start=1):
                    cell = ws.cell(row=1, column=col_idx, value=field)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align

                even_fill = PatternFill("solid", fgColor="D9E1F2")
                odd_fill  = PatternFill("solid", fgColor="FFFFFF")

                for row_idx, row_data in enumerate(data_list[1:], start=2):
                    fill = even_fill if row_idx % 2 == 0 else odd_fill
                    for col_idx, value in enumerate(row_data, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.fill = fill
                        cell.alignment = Alignment(vertical="center")

                for col in ws.columns:
                    max_len = max((len(str(c.value)) if c.value else 0) for c in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

                ws.row_dimensions[1].height = 40
                ws.freeze_panes = "A2"

                wb.save(destination)
                rows_count = len(data_list) - 1
                self.after(0, lambda: self._on_success(rows_count, destination))
            else:
                self.after(0, lambda: self._on_no_data())

        except Exception as exc:
            err_msg = str(exc)
            self.after(0, lambda: self._on_error(err_msg))

    # ── Completion Callbacks (executed on main thread) ─────────────────────
    def _on_success(self, rows: int, destination: str):
        self.is_running = False
        self.progressbar.stop()
        self.progressbar.config(
            style="done.Horizontal.TProgressbar",
            mode="determinate", value=100
        )
        self.status_lbl.config(
            text=f"✅  Completed — {rows} row(s) exported → {Path(destination).name}",
            fg=OK_GREEN
        )
        self._set_controls(locked=False)
        messagebox.showinfo(
            "Extraction Completed",
            f"{rows} row(s) were exported to the file:\n{destination}"
        )

    def _on_no_data(self):
        self.is_running = False
        self.progressbar.stop()
        self.progressbar.config(
            style="idle.Horizontal.TProgressbar",
            mode="determinate", value=0
        )
        self.status_lbl.config(
            text="⚠️  No data found to export.",
            fg="#E65100"
        )
        self._set_controls(locked=False)
        messagebox.showwarning(
            "No Data",
            "No valid SGI files were found in the selected folder."
        )

    def _on_error(self, msg: str):
        self.is_running = False
        self.progressbar.stop()
        self.progressbar.config(
            style="idle.Horizontal.TProgressbar",
            mode="determinate", value=0
        )
        self.status_lbl.config(text="❌  Error during extraction.", fg="#C62828")
        self._set_controls(locked=False)
        messagebox.showerror(
            "Extraction Error",
            f"An unexpected error occurred:\n\n{msg}"
        )


# ────────────────────────────────────────────────────────────────────────────
# Entry Point
# ────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = SGIApp()
    app.mainloop()