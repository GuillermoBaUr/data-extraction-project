"""
interfaz_sgi.py
---------------
Interfaz gráfica para el extractor de datos SGI.
Ejecutar desde la misma carpeta que extraer_sgi.py.

Requiere:
    pip install python-docx openpyxl pillow
"""

import os
import sys
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path

# ── Intentar cargar Pillow para mostrar el logo PNG ──────────────────────────
try:
    from PIL import Image, ImageTk
    PILLOW_OK = True
except ImportError:
    PILLOW_OK = False


# ────────────────────────────────────────────────────────────────────────────
# Paleta de colores
# ────────────────────────────────────────────────────────────────────────────
AZUL_HEADER   = "#0D3B6E"   # Azul oscuro institucional
AZUL_BOTON    = "#1565C0"   # Azul botones activos
AZUL_HOVER    = "#1976D2"   # Hover de botones
GRIS_DISABLED = "#90A4AE"   # Botón deshabilitado
VERDE_OK      = "#2E7D32"   # Barra completado
GRIS_RUNNING  = "#546E7A"   # Barra en ejecución
BLANCO        = "#FFFFFF"
FONDO_BODY    = "#FFFFFF"
FONDO_CAMPO   = "#F5F7FA"
BORDE_CAMPO   = "#CFD8DC"
TEXTO_DARK    = "#1A237E"
TEXTO_LABEL   = "#37474F"
SUBHEADER_BG  = "#E3EBF6"


# ────────────────────────────────────────────────────────────────────────────
# Clase principal
# ────────────────────────────────────────────────────────────────────────────
class SGIApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("SGC Extractor de Datos")
        self.resizable(False, False)
        self.configure(bg=FONDO_BODY)

        # Intentar cargar icono .ico
        ico_path = Path("../img/tecnm.ico")
        if ico_path.exists():
            try:
                self.iconbitmap(str(ico_path))
            except Exception:
                pass

        # Variables
        self.var_carpeta_origen = tk.StringVar()
        self.var_archivo_destino = tk.StringVar()
        self.running = False

        self._build_ui()
        self._center_window()

    # ── Centrar ventana en pantalla ─────────────────────────────────────────
    def _center_window(self):
        self.update_idletasks()
        w = self.winfo_width()
        h = self.winfo_height()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.geometry(f"+{x}+{y}")

    # ── Construcción de la UI ───────────────────────────────────────────────
    def _build_ui(self):
        self._build_header()
        self._build_subheader()
        self._build_body()

    # ── Header azul con logo ────────────────────────────────────────────────
    def _build_header(self):
        header = tk.Frame(self, bg=AZUL_HEADER, height=80)
        header.pack(fill="x")
        header.pack_propagate(False)

        # Logo PNG a la izquierda
        logo_path = Path("../img/tecnm.png")
        if PILLOW_OK and logo_path.exists():
            try:
                img = Image.open(logo_path).convert("RGBA")
                img.thumbnail((60, 60), Image.LANCZOS)
                self._logo_img = ImageTk.PhotoImage(img)
                tk.Label(
                    header, image=self._logo_img,
                    bg=AZUL_HEADER, bd=0
                ).pack(side="left", padx=(18, 10), pady=10)
            except Exception:
                self._logo_img = None
        else:
            # Cuadrado de placeholder si no hay imagen
            placeholder = tk.Label(
                header, text="TecNM", bg=BLANCO, fg=AZUL_HEADER,
                font=("Georgia", 9, "bold"), width=7, height=3, relief="flat"
            )
            placeholder.pack(side="left", padx=(18, 10), pady=10)

        # Texto del header
        text_frame = tk.Frame(header, bg=AZUL_HEADER)
        text_frame.pack(side="left", fill="both", expand=True, pady=10)

        tk.Label(
            text_frame,
            text="Tecnológico Nacional de México",
            font=("Georgia", 15, "bold"),
            bg=AZUL_HEADER, fg=BLANCO, anchor="w"
        ).pack(anchor="w")

        tk.Label(
            text_frame,
            text="Campus San Juan del Río",
            font=("Georgia", 11),
            bg=AZUL_HEADER, fg="#BBDEFB", anchor="w"
        ).pack(anchor="w")

    # ── Subheader institucional ─────────────────────────────────────────────
    def _build_subheader(self):
        sub = tk.Frame(self, bg=SUBHEADER_BG, pady=8)
        sub.pack(fill="x")

        tk.Label(
            sub,
            text="Departamento de la Carrera de Ingeniería Industrial",
            font=("Segoe UI", 10, "bold"),
            bg=SUBHEADER_BG, fg=TEXTO_DARK
        ).pack()

        tk.Label(
            sub,
            text="Supervisora:  María Patricia Uribe Rodríguez",
            font=("Segoe UI", 9),
            bg=SUBHEADER_BG, fg=TEXTO_LABEL
        ).pack()

    # ── Cuerpo principal ────────────────────────────────────────────────────
    def _build_body(self):
        body = tk.Frame(self, bg=FONDO_BODY, padx=32, pady=24)
        body.pack(fill="both", expand=True)

        # Título del programa
        tk.Label(
            body,
            text="SGC  —  Extractor de Datos",
            font=("Georgia", 14, "bold"),
            bg=FONDO_BODY, fg=TEXTO_DARK
        ).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 20))

        # ── Selector de carpeta origen ──────────────────────────────────────
        self._build_field_row(
            body, row=1,
            label="Carpeta de archivos ZIP / DOCX:",
            variable=self.var_carpeta_origen,
            btn_text="📂  Seleccionar carpeta",
            cmd=self._seleccionar_carpeta,
            placeholder="Selecciona la carpeta que contiene los archivos…"
        )

        # ── Selector de archivo destino Excel ──────────────────────────────
        self._build_field_row(
            body, row=3,
            label="Guardar Excel en:",
            variable=self.var_archivo_destino,
            btn_text="💾  Guardar como…",
            cmd=self._seleccionar_destino,
            placeholder="Elige dónde guardar el archivo datos.xlsx…"
        )

        # Separador
        sep = tk.Frame(body, bg=BORDE_CAMPO, height=1)
        sep.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(20, 0))

        # ── Botón Iniciar ───────────────────────────────────────────────────
        self.btn_iniciar = tk.Button(
            body,
            text="▶   Iniciar Extracción",
            font=("Segoe UI", 11, "bold"),
            bg=GRIS_DISABLED, fg=BLANCO,
            activebackground=AZUL_HOVER, activeforeground=BLANCO,
            relief="flat", cursor="arrow",
            padx=28, pady=10,
            state="disabled",
            command=self._iniciar_extraccion
        )
        self.btn_iniciar.grid(row=6, column=0, columnspan=3, pady=(18, 0))

        # ── Barra de estado ─────────────────────────────────────────────────
        status_frame = tk.Frame(body, bg=FONDO_BODY)
        status_frame.grid(row=7, column=0, columnspan=3, sticky="ew", pady=(20, 0))
        status_frame.columnconfigure(0, weight=1)

        self.lbl_status = tk.Label(
            status_frame,
            text="",
            font=("Segoe UI", 9),
            bg=FONDO_BODY, fg=TEXTO_LABEL, anchor="w"
        )
        self.lbl_status.grid(row=0, column=0, sticky="w", pady=(0, 4))

        # Barra de progreso (estilo propio)
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure(
            "idle.Horizontal.TProgressbar",
            troughcolor=BORDE_CAMPO, background=BORDE_CAMPO,
            thickness=14, borderwidth=0
        )
        style.configure(
            "running.Horizontal.TProgressbar",
            troughcolor=BORDE_CAMPO, background=GRIS_RUNNING,
            thickness=14, borderwidth=0
        )
        style.configure(
            "done.Horizontal.TProgressbar",
            troughcolor=BORDE_CAMPO, background=VERDE_OK,
            thickness=14, borderwidth=0
        )

        self.progressbar = ttk.Progressbar(
            status_frame,
            style="idle.Horizontal.TProgressbar",
            mode="indeterminate", length=460
        )
        self.progressbar.grid(row=1, column=0, sticky="ew")

        body.columnconfigure(1, weight=1)
        body.config(width=560)

    # ── Helper: fila etiqueta + entry + botón ──────────────────────────────
    def _build_field_row(self, parent, row, label, variable, btn_text, cmd, placeholder):
        tk.Label(
            parent, text=label,
            font=("Segoe UI", 9, "bold"),
            bg=FONDO_BODY, fg=TEXTO_LABEL, anchor="w"
        ).grid(row=row, column=0, columnspan=3, sticky="w", pady=(0, 4))

        entry_frame = tk.Frame(parent, bg=BORDE_CAMPO, bd=0)
        entry_frame.grid(row=row+1, column=0, columnspan=2, sticky="ew", ipady=1, ipadx=1)
        entry_frame.columnconfigure(0, weight=1)

        entry = tk.Entry(
            entry_frame,
            textvariable=variable,
            font=("Segoe UI", 9),
            bg=FONDO_CAMPO, fg=TEXTO_LABEL,
            relief="flat", state="readonly",
            readonlybackground=FONDO_CAMPO,
            bd=6, width=48
        )
        entry.grid(row=0, column=0, sticky="ew")

        # Placeholder visual
        if not variable.get():
            entry.config(state="normal")
            entry.insert(0, placeholder)
            entry.config(fg="#B0BEC5", state="readonly")

        def _on_var_change(*_):
            val = variable.get()
            entry.config(state="normal")
            entry.delete(0, "end")
            if val:
                entry.insert(0, val)
                entry.config(fg=TEXTO_LABEL)
            else:
                entry.insert(0, placeholder)
                entry.config(fg="#B0BEC5")
            entry.config(state="readonly")
            self._actualizar_boton_iniciar()

        variable.trace_add("write", _on_var_change)

        # Guardar referencia al entry para poder bloquearlo
        attr = "entry_carpeta" if "carpeta" in label.lower() else "entry_destino"
        setattr(self, attr, entry)

        btn = tk.Button(
            parent, text=btn_text,
            font=("Segoe UI", 9),
            bg=AZUL_BOTON, fg=BLANCO,
            activebackground=AZUL_HOVER, activeforeground=BLANCO,
            relief="flat", cursor="hand2",
            padx=12, pady=6,
            command=cmd
        )
        btn.grid(row=row+1, column=2, padx=(10, 0), sticky="w")
        self._add_hover(btn, AZUL_BOTON, AZUL_HOVER)

        # Guardar referencia al botón
        attr_btn = "btn_carpeta" if "carpeta" in label.lower() else "btn_destino"
        setattr(self, attr_btn, btn)

        parent.rowconfigure(row+2, minsize=12)

    # ── Hover helper ────────────────────────────────────────────────────────
    @staticmethod
    def _add_hover(widget, normal_color, hover_color):
        widget.bind("<Enter>", lambda e: widget.config(bg=hover_color))
        widget.bind("<Leave>", lambda e: widget.config(bg=normal_color))

    # ── Selección de carpeta origen ─────────────────────────────────────────
    def _seleccionar_carpeta(self):
        carpeta = filedialog.askdirectory(
            title="Seleccionar carpeta con archivos ZIP / DOCX",
            mustexist=True
        )
        if carpeta:
            self.var_carpeta_origen.set(carpeta)

    # ── Selección de archivo destino ────────────────────────────────────────
    def _seleccionar_destino(self):
        archivo = filedialog.asksaveasfilename(
            title="Guardar Excel como…",
            defaultextension=".xlsx",
            filetypes=[("Archivo Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
            initialfile="datos.xlsx"
        )
        if archivo:
            self.var_archivo_destino.set(archivo)

    # ── Habilitar / deshabilitar botón Iniciar ──────────────────────────────
    def _actualizar_boton_iniciar(self):
        if self.running:
            return
        if self.var_carpeta_origen.get() and self.var_archivo_destino.get():
            self.btn_iniciar.config(
                state="normal", bg=AZUL_BOTON, cursor="hand2"
            )
            self._add_hover(self.btn_iniciar, AZUL_BOTON, AZUL_HOVER)
        else:
            self.btn_iniciar.config(
                state="disabled", bg=GRIS_DISABLED, cursor="arrow"
            )

    # ── Bloquear / desbloquear controles ────────────────────────────────────
    def _set_controls(self, locked: bool):
        state = "disabled" if locked else "normal"
        self.btn_carpeta.config(state=state)
        self.btn_destino.config(state=state)
        if not locked:
            self._actualizar_boton_iniciar()
        else:
            self.btn_iniciar.config(state="disabled", bg=GRIS_DISABLED, cursor="arrow")

    # ── Iniciar extracción en hilo separado ─────────────────────────────────
    def _iniciar_extraccion(self):
        if self.running:
            return
        self.running = True
        self._set_controls(locked=True)

        # Estado: en ejecución
        self.progressbar.config(style="running.Horizontal.TProgressbar", mode="indeterminate")
        self.progressbar.start(12)
        self.lbl_status.config(text="⏳  Extrayendo datos, por favor espera…", fg=GRIS_RUNNING)

        hilo = threading.Thread(target=self._ejecutar_extraccion, daemon=True)
        hilo.start()

    def _ejecutar_extraccion(self):
        """Corre extraer_sgi en un hilo para no congelar la UI."""
        try:
            # Importar el módulo de extracción (debe estar en la misma carpeta)
            import importlib.util, types

            spec = importlib.util.spec_from_file_location(
                "extraer_sgi",
                Path(__file__).parent / "extraer_sgi.py"
            )
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)

            carpeta = self.var_carpeta_origen.get()
            destino = self.var_archivo_destino.get()

            # ── Capturar stdout para no romper la UI ─────────────────────
            import io, contextlib
            buffer = io.StringIO()
            data_list = [mod.ENCABEZADO_EXCEL]

            with contextlib.redirect_stdout(buffer):
                mod.procesar_ruta(carpeta, data_list)

            log_output = buffer.getvalue()

            if len(data_list) > 1:
                # Guardar en la ruta elegida por el usuario
                os.chdir(str(Path(destino).parent))
                mod.crear_excel.__globals__  # acceso al módulo original

                # Reescribimos el Excel en la ruta correcta
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "SGI Datos"

                fill_header = PatternFill("solid", fgColor="1F4E79")
                font_header = Font(bold=True, color="FFFFFF", size=11)
                alin_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

                for col_idx, campo in enumerate(data_list[0], start=1):
                    cell = ws.cell(row=1, column=col_idx, value=campo)
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = alin_center

                fill_par   = PatternFill("solid", fgColor="D9E1F2")
                fill_impar = PatternFill("solid", fgColor="FFFFFF")

                for row_idx, fila in enumerate(data_list[1:], start=2):
                    fill = fill_par if row_idx % 2 == 0 else fill_impar
                    for col_idx, valor in enumerate(fila, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                        cell.fill = fill
                        cell.alignment = Alignment(vertical="center")

                for col in ws.columns:
                    max_len = max((len(str(c.value)) if c.value else 0) for c in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

                ws.row_dimensions[1].height = 40
                ws.freeze_panes = "A2"

                wb.save(destino)
                filas = len(data_list) - 1
                self.after(0, lambda: self._on_success(filas, destino))
            else:
                self.after(0, lambda: self._on_no_data())

        except Exception as exc:
            msg = str(exc)
            self.after(0, lambda: self._on_error(msg))

    # ── Callbacks de finalización (ejecutados en el hilo principal) ─────────
    def _on_success(self, filas: int, destino: str):
        self.running = False
        self.progressbar.stop()
        self.progressbar.config(
            style="done.Horizontal.TProgressbar",
            mode="determinate", value=100
        )
        self.lbl_status.config(
            text=f"✅  Completado — {filas} fila(s) exportada(s) → {Path(destino).name}",
            fg=VERDE_OK
        )
        self._set_controls(locked=False)
        messagebox.showinfo(
            "Extracción completada",
            f"Se exportaron {filas} fila(s) al archivo:\n{destino}"
        )

    def _on_no_data(self):
        self.running = False
        self.progressbar.stop()
        self.progressbar.config(
            style="idle.Horizontal.TProgressbar",
            mode="determinate", value=0
        )
        self.lbl_status.config(
            text="⚠️  No se encontraron datos para exportar.",
            fg="#E65100"
        )
        self._set_controls(locked=False)
        messagebox.showwarning(
            "Sin datos",
            "No se encontraron archivos SGI válidos en la carpeta seleccionada."
        )

    def _on_error(self, msg: str):
        self.running = False
        self.progressbar.stop()
        self.progressbar.config(
            style="idle.Horizontal.TProgressbar",
            mode="determinate", value=0
        )
        self.lbl_status.config(text="❌  Error durante la extracción.", fg="#C62828")
        self._set_controls(locked=False)
        messagebox.showerror(
            "Error en la extracción",
            f"Ocurrió un error inesperado:\n\n{msg}"
        )


# ────────────────────────────────────────────────────────────────────────────
# Punto de entrada
# ────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = SGIApp()
    app.mainloop()